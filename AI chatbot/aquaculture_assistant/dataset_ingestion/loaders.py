"""Load CSV, JSON, text, Markdown, PDF, and Excel knowledge sources."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

SUPPORTED_EXTENSIONS = {
    ".csv",
    ".json",
    ".txt",
    ".md",
    ".markdown",
    ".pdf",
    ".xlsx",
    ".xls",
}
IGNORED_DIRECTORY_NAMES = {
    ".git",
    ".venv",
    "__pycache__",
    "models",
    "vector_db",
    "chat_history",
    "build",
}


@dataclass(frozen=True)
class SourceDocument:
    text: str
    source: str
    title: str


def discover_source_files(roots: Iterable[Path]) -> list[Path]:
    """Find supported documents deterministically without following artifacts."""
    found: set[Path] = set()
    for root in roots:
        root = Path(root)
        if root.is_file() and root.suffix.lower() in SUPPORTED_EXTENSIONS:
            found.add(root.resolve())
            continue
        if not root.is_dir():
            continue
        for path in root.rglob("*"):
            if any(
                part.lower() in IGNORED_DIRECTORY_NAMES for part in path.parts
            ):
                continue
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
                found.add(path.resolve())
    return sorted(found, key=lambda path: str(path).lower())


def _humanize_json(value: Any, prefix: str = "") -> str:
    lines: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            label = str(key).replace("_", " ").strip().title()
            next_prefix = f"{prefix} {label}".strip()
            if isinstance(item, (dict, list)):
                nested = _humanize_json(item, next_prefix)
                if nested:
                    lines.append(nested)
            elif item not in (None, ""):
                lines.append(f"{next_prefix}: {item}")
    elif isinstance(value, list):
        for index, item in enumerate(value, start=1):
            nested = _humanize_json(item, f"{prefix} Item {index}".strip())
            if nested:
                lines.append(nested)
    elif value not in (None, ""):
        lines.append(f"{prefix}: {value}" if prefix else str(value))
    return "\n".join(lines)


def _load_json(path: Path) -> list[SourceDocument]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload if isinstance(payload, list) else [payload]
    documents: list[SourceDocument] = []
    for index, item in enumerate(items, start=1):
        if isinstance(item, dict):
            title = str(
                item.get("name")
                or item.get("disease_name")
                or item.get("title")
                or f"{path.stem} item {index}"
            )
        else:
            title = f"{path.stem} item {index}"
        text = _humanize_json(item).strip()
        if text:
            documents.append(SourceDocument(text, str(path), title))
    return documents


def _load_csv(path: Path) -> list[SourceDocument]:
    documents: list[SourceDocument] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader, start=1):
            values = [
                f"{key.replace('_', ' ').title()}: {value.strip()}"
                for key, value in row.items()
                if key and value and value.strip()
            ]
            if values:
                documents.append(
                    SourceDocument(
                        text="\n".join(values),
                        source=str(path),
                        title=f"{path.stem} row {index}",
                    )
                )
    return documents


def _load_pdf(path: Path) -> list[SourceDocument]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError(
            "PDF ingestion requires pypdf. Install requirements-assistant.txt."
        ) from exc

    documents: list[SourceDocument] = []
    reader = PdfReader(str(path))
    for index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            documents.append(
                SourceDocument(
                    text=text,
                    source=str(path),
                    title=f"{path.stem}, page {index}",
                )
            )
    return documents


def _load_excel(path: Path) -> list[SourceDocument]:
    if path.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "Legacy Excel ingestion requires xlrd. "
                "Install requirements-assistant.txt."
            ) from exc
        workbook = xlrd.open_workbook(path)
        documents: list[SourceDocument] = []
        for sheet in workbook.sheets():
            if sheet.nrows == 0:
                continue
            headers = sheet.row_values(0)
            for row_index in range(1, sheet.nrows):
                row = sheet.row_values(row_index)
                values = []
                for index, value in enumerate(row):
                    if value in (None, ""):
                        continue
                    header = (
                        str(headers[index])
                        if index < len(headers) and headers[index] not in (None, "")
                        else f"Column {index + 1}"
                    )
                    values.append(f"{header}: {value}")
                if values:
                    documents.append(
                        SourceDocument(
                            text="\n".join(values),
                            source=str(path),
                            title=(
                                f"{path.stem}/{sheet.name} row {row_index + 1}"
                            ),
                        )
                    )
        return documents
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel ingestion requires openpyxl. "
            "Install requirements-assistant.txt."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    documents: list[SourceDocument] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        for row_number, row in enumerate(rows, start=2):
            values = []
            for index, value in enumerate(row):
                if value in (None, ""):
                    continue
                header = (
                    str(headers[index])
                    if index < len(headers) and headers[index] not in (None, "")
                    else f"Column {index + 1}"
                )
                values.append(f"{header}: {value}")
            if values:
                documents.append(
                    SourceDocument(
                        text="\n".join(values),
                        source=str(path),
                        title=f"{path.stem}/{sheet.title} row {row_number}",
                    )
                )
    workbook.close()
    return documents


def load_source_documents(path: Path) -> list[SourceDocument]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".csv":
        return _load_csv(path)
    if suffix == ".pdf":
        return _load_pdf(path)
    if suffix in {".xlsx", ".xls"}:
        return _load_excel(path)
    if suffix in {".txt", ".md", ".markdown"}:
        text = path.read_text(encoding="utf-8").strip()
        return [SourceDocument(text, str(path), path.stem)] if text else []
    return []
